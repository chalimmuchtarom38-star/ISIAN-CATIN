import os
import openpyxl
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta, date
from io import BytesIO
from reportlab.lib.pagesizes import portrait
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

st.set_page_config(page_title="Form Input Catin F4", layout="wide")
st.title("📝 Formulir Input Isian Data Catin (F4)")

FILE_NAME = "BERKAS_CATIN_F4.xlsx"

if not os.path.exists(FILE_NAME):
    st.error(f"File '{FILE_NAME}' tidak ditemukan! Silakan simpan file master ke format .xlsx dengan nama BERKAS_CATIN_F4.xlsx di folder yang sama.")
    st.stop()

# Daftar Nama Bulan Bahasa Indonesia
NAMA_BULAN = [
    "", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"
]

def parse_to_date_object(val):
    """Mencoba mengubah string/excel serial/date menjadi object date Python untuk st.date_input"""
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    if pd.isna(val) or val is None or str(val).strip() == "":
        return date.today()
    if isinstance(val, (int, float)):
        try:
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
            return dt.date()
        except:
            return date.today()
    
    val_str = str(val).strip()
    if val_str.isdigit() and len(val_str) >= 5:
        try:
            dt = datetime(1899, 12, 30) + timedelta(days=int(val_str))
            return dt.date()
        except:
            pass
            
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
            
    return date.today()

def format_tanggal_indonesia(val):
    """Mengubah date object atau nilai tanggal menjadi string 'DD BULAN YYYY'"""
    if pd.isna(val) or val is None or str(val).strip() == "":
        return ""
    if isinstance(val, (date, datetime)):
        return f"{val.day} {NAMA_BULAN[val.month]} {val.year}"
    if isinstance(val, (int, float)):
        try:
            date_obj = datetime(1899, 12, 30) + timedelta(days=int(val))
            return f"{date_obj.day} {NAMA_BULAN[date_obj.month]} {date_obj.year}"
        except:
            return str(val)
    val_str = str(val).strip()
    if val_str.isdigit() and len(val_str) >= 5:
        try:
            date_obj = datetime(1899, 12, 30) + timedelta(days=int(val_str))
            return f"{date_obj.day} {NAMA_BULAN[date_obj.month]} {date_obj.year}"
        except:
            return val_str
    return val_str

# Pemetaan Baris Excel (1-based row index)
MAPPING_ROWS = [
    ("Baris 2", "Nomor Register", 2),
    ("Baris 3", "Tanggal Surat", 3),
    ("Baris 4", "Tanggal Pelaksanaan", 4),
    ("Baris 5", "Jam Pelaksanaan", 5),
    ("Baris 6", "Tempat Akad Nikah", 6),
    
    # Catin Laki-Laki
    ("Baris 8", "Nama Catin Laki-Laki", 8),
    ("Baris 9", "Bin (Ayah Laki-Laki)", 9),
    ("Baris 10", "TTL Catin Laki-Laki", 10),
    ("Baris 11", "NIK Catin Laki-Laki", 11),
    ("Baris 12", "Pekerjaan Laki-Laki", 12),
    ("Baris 13", "Status Laki-Laki", 13),
    ("Baris 14", "Jenis Kelamin Laki-Laki", 14),
    ("Baris 15", "Nama Istri Terdahulu", 15),
    ("Baris 16", "Alamat Catin Laki-Laki", 16),
    ("Baris 17", "Pendidikan Laki-Laki", 17),
    
    # Ortu Laki-Laki
    ("Baris 19", "Nama Ayah Laki-Laki", 19),
    ("Baris 20", "NIK Ayah Laki-Laki", 20),
    ("Baris 21", "TTL Ayah Laki-Laki", 21),
    ("Baris 22", "Pekerjaan Ayah Laki-Laki", 22),
    ("Baris 23", "Alamat Ayah Laki-Laki", 23),
    ("Baris 26", "Nama Ibu Laki-Laki", 26),
    ("Baris 27", "NIK Ibu Laki-Laki", 27),
    ("Baris 28", "TTL Ibu Laki-Laki", 28),
    ("Baris 29", "Pekerjaan Ibu Laki-Laki", 29),
    ("Baris 30", "Alamat Ibu Laki-Laki", 30),
    
    # Catin Perempuan
    ("Baris 34", "Nama Catin Perempuan", 34),
    ("Baris 35", "Binti (Ayah Perempuan)", 35),
    ("Baris 36", "TTL Catin Perempuan", 36),
    ("Baris 37", "NIK Catin Perempuan", 37),
    ("Baris 38", "Pekerjaan Perempuan", 38),
    ("Baris 39", "Status Perempuan", 39),
    ("Baris 40", "Jenis Kelamin Perempuan", 40),
    ("Baris 41", "Alamat Catin Perempuan", 41),
    ("Baris 42", "Nama Suami Terdahulu", 42),
    ("Baris 43", "Pendidikan Perempuan", 43),
    
    # Ortu Perempuan
    ("Baris 45", "Nama Ayah Perempuan", 45),
    ("Baris 46", "NIK Ayah Perempuan", 46),
    ("Baris 47", "TTL Ayah Perempuan", 47),
    ("Baris 48", "Pekerjaan Ayah Perempuan", 48),
    ("Baris 49", "Alamat Ayah Perempuan", 49),
    ("Baris 52", "Nama Ibu Perempuan", 52),
    ("Baris 53", "NIK Ibu Perempuan", 53),
    ("Baris 54", "TTL Ibu Perempuan", 54),
    ("Baris 55", "Pekerjaan Ibu Perempuan", 55),
    ("Baris 56", "Alamat Ibu Perempuan", 56),
    
    # Wali & Saksi
    ("Baris 58", "Nama Wali", 58),
    ("Baris 59", "Bin Wali", 59),
    ("Baris 60", "NIK Wali", 60),
    ("Baris 61", "TTL Wali", 61),
    ("Baris 62", "Pekerjaan Wali", 62),
    ("Baris 63", "Alamat Wali", 63),
    ("Baris 64", "Hubungan Wali", 64),
    ("Baris 65", "Mahar / Maskawin", 65),
    ("Baris 70", "Nama Saksi 1", 70),
    ("Baris 71", "TTL Saksi 1", 71),
    ("Baris 72", "NIK Saksi 1", 72),
    ("Baris 73", "Pekerjaan Saksi 1", 73),
    ("Baris 74", "Alamat Saksi 1", 74),
    ("Baris 76", "Nama Saksi 2", 76),
    ("Baris 77", "TTL Saksi 2", 77),
    ("Baris 78", "NIK Saksi 2", 78),
    ("Baris 79", "Pekerjaan Saksi 2", 79),
    ("Baris 80", "Alamat Saksi 2", 80)
]

@st.cache_data(ttl=1)
def load_excel_data():
    try:
        wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
        sheet_name = next((s for s in wb.sheetnames if s.strip().upper() == "ISIAN DATA"), wb.sheetnames[0])
        ws = wb[sheet_name]
        
        extracted_data = []
        for ref, label, r_idx in MAPPING_ROWS:
            val = ws.cell(row=r_idx, column=7).value
            if val is None:
                val = ws.cell(row=r_idx, column=6).value
            
            if label not in ["Tanggal Surat", "Tanggal Pelaksanaan"]:
                val = format_tanggal_indonesia(val)
                if str(val).strip() == ":":
                    val = ""
            extracted_data.append((ref, label, val, r_idx))
            
        return extracted_data
    except Exception as e:
        st.error(f"Gagal membaca file Excel: {e}")
        return []

data_list = load_excel_data()

# FORM INPUT INTERAKTIF STREAMLIT
with st.form("form_catin"):
    st.subheader("Isi / Sesuaikan Data Catin Secara Praktis")
    
    t1, t2, t3, t4 = st.tabs([
        "📄 Register & Surat", 
        "👨 Catin Laki-Laki", 
        "👩 Catin Perempuan", 
        "🤝 Wali & Saksi"
    ])
    
    user_inputs = {}

    with t1:
        st.markdown("##### Data Surat & Pelaksanaan Akad")
        for ref, label, val, row_num in data_list[:5]:
            if label in ["Tanggal Surat", "Tanggal Pelaksanaan"]:
                default_dt = parse_to_date_object(val)
                d_selected = st.date_input(
                    f"[{ref}] {label}", 
                    value=default_dt, 
                    format="DD/MM/YYYY",
                    key=f"inp_{row_num}"
                )
                user_inputs[label] = format_tanggal_indonesia(d_selected)
            else:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")

    with t2:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("##### -- Data Catin Laki-Laki --")
            for ref, label, val, row_num in data_list[5:15]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")
        with col2:
            st.markdown("##### -- Data Orang Tua Laki-Laki --")
            for ref, label, val, row_num in data_list[15:25]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")

    with t3:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("##### -- Data Catin Perempuan --")
            for ref, label, val, row_num in data_list[25:35]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")
        with col2:
            st.markdown("##### -- Data Orang Tua Perempuan --")
            for ref, label, val, row_num in data_list[35:45]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")

    with t4:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("##### -- Data Wali & Mahar --")
            for ref, label, val, row_num in data_list[45:53]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")
        with col2:
            st.markdown("##### -- Data Saksi 1 & Saksi 2 --")
            for ref, label, val, row_num in data_list[53:]:
                user_inputs[label] = st.text_input(f"[{ref}] {label}", value=str(val or ""), key=f"inp_{row_num}")

    btn_simpan = st.form_submit_button("💾 Simpan & Perbarui Data Input", use_container_width=True)

if btn_simpan:
    st.session_state['input_data'] = user_inputs
    st.success("✅ Data berhasil diperbarui di memori sistem!")

# PANEL DOWNLOAD RESULT
st.markdown("---")
st.subheader("📥 Download Hasil Isian")

col_d1, col_d2 = st.columns(2)

# FUNGSI MENGEDIT FILE EXCEL MASTER .XLSX SECARA UTUH (TIDAK DIUTAK-ATIK)
def generate_updated_xlsx(data_dict):
    output = BytesIO()
    wb = openpyxl.load_workbook(FILE_NAME)
    
    sheet_name = next((s for s in wb.sheetnames if s.strip().upper() == "ISIAN DATA"), wb.sheetnames[0])
    ws = wb[sheet_name]
    
    current_inputs = data_dict if data_dict else {
        item[1]: format_tanggal_indonesia(item[2]) if item[1] in ["Tanggal Surat", "Tanggal Pelaksanaan"] else item[2] 
        for item in data_list
    }
    label_to_row = {item[1]: item[2] for item in MAPPING_ROWS}
    
    for label, val in current_inputs.items():
        if label in label_to_row:
            r_idx = label_to_row[label]
            ws.cell(row=r_idx, column=7, value=val)

    wb.save(output)
    output.seek(0)
    return output

with col_d1:
    excel_bytes = generate_updated_xlsx(st.session_state.get('input_data'))
    
    curr_data = st.session_state.get('input_data', {
        item[1]: format_tanggal_indonesia(item[2]) if item[1] in ["Tanggal Surat", "Tanggal Pelaksanaan"] else item[2] 
        for item in data_list
    })
    pria = str(curr_data.get("Nama Catin Laki-Laki", "")).strip().upper()
    wanita = str(curr_data.get("Nama Catin Perempuan", "")).strip().upper()
    file_excel_name = f"BERKAS_CATIN_{pria}_&_{wanita}.xlsx" if (pria or wanita) else "BERKAS_CATIN_TERISI.xlsx"

    st.download_button(
        label="📊 Download File Excel Terisi Utuh (.xlsx)",
        data=excel_bytes,
        file_name=file_excel_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with col_d2:
    # FUNGSI PDF DIPERBARUI: FORMAT PORTRAIT F4 FULL 1 LEMBAR DENGAN BLOK TANDA TANGAN KASI & KEPALA DESA TAMBI
    def generate_pdf_formal(data_dict):
        buffer = BytesIO()
        F4_PORTRAIT = portrait((612, 936))
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=F4_PORTRAIT, 
            rightMargin=20, 
            leftMargin=20, 
            topMargin=15, 
            bottomMargin=15
        )
        elements = []
        styles = getSampleStyleSheet()
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading1'],
            fontSize=11,
            leading=13,
            alignment=1,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor("#0F2C59")
        )
        
        sub_header_style = ParagraphStyle(
            'SubHeaderStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=1,
            fontName='Helvetica-Oblique',
            textColor=colors.HexColor("#333333")
        )
        
        sec_header = ParagraphStyle('SecHeader', fontSize=7, leading=8, fontName='Helvetica-Bold', textColor=colors.whitesmoke)
        cell_bold = ParagraphStyle('CB', fontSize=6.5, leading=7.5, fontName='Helvetica-Bold')
        cell_norm = ParagraphStyle('CN', fontSize=6.5, leading=7.5, fontName='Helvetica')
        
        # Style Tanda Tangan
        ttd_center = ParagraphStyle('TTDCenter', fontSize=7.5, leading=9, alignment=1, fontName='Helvetica')
        ttd_bold_underline = ParagraphStyle('TTDBold', fontSize=8, leading=10, alignment=1, fontName='Helvetica-Bold')

        elements.append(Paragraph("RINGKASAN ISIAN DATA BERKAS CATIN (F4)", header_style))
        elements.append(Paragraph("Daftar Pemeriksaan & Verifikasi Data Pernikahan Terkelompok", sub_header_style))
        elements.append(Spacer(1, 4))
        
        current_data = data_dict if data_dict else {
            item[1]: format_tanggal_indonesia(item[2]) if item[1] in ["Tanggal Surat", "Tanggal Pelaksanaan"] else item[2] 
            for item in data_list
        }
        
        groups = [
            ("I. REGISTER & PELAKSANAAN AKAD", [
                "Nomor Register", "Tanggal Surat", "Tanggal Pelaksanaan", "Jam Pelaksanaan", "Tempat Akad Nikah"
            ]),
            ("II. DATA CATIN LAKI-LAKI", [
                "Nama Catin Laki-Laki", "Bin (Ayah Laki-Laki)", "TTL Catin Laki-Laki", "NIK Catin Laki-Laki", 
                "Pekerjaan Laki-Laki", "Status Laki-Laki", "Jenis Kelamin Laki-Laki", "Nama Istri Terdahulu", 
                "Alamat Catin Laki-Laki", "Pendidikan Laki-Laki"
            ]),
            ("III. DATA ORANG TUA LAKI-LAKI", [
                "Nama Ayah Laki-Laki", "NIK Ayah Laki-Laki", "TTL Ayah Laki-Laki", "Pekerjaan Ayah Laki-Laki", "Alamat Ayah Laki-Laki",
                "Nama Ibu Laki-Laki", "NIK Ibu Laki-Laki", "TTL Ibu Laki-Laki", "Pekerjaan Ibu Laki-Laki", "Alamat Ibu Laki-Laki"
            ]),
            ("IV. DATA CATIN PEREMPUAN", [
                "Nama Catin Perempuan", "Binti (Ayah Perempuan)", "TTL Catin Perempuan", "NIK Catin Perempuan", 
                "Pekerjaan Perempuan", "Status Perempuan", "Jenis Kelamin Perempuan", "Alamat Catin Perempuan", 
                "Nama Suami Terdahulu", "Pendidikan Perempuan"
            ]),
            ("V. DATA ORANG TUA PEREMPUAN", [
                "Nama Ayah Perempuan", "NIK Ayah Perempuan", "TTL Ayah Perempuan", "Pekerjaan Ayah Perempuan", "Alamat Ayah Perempuan",
                "Nama Ibu Perempuan", "NIK Ibu Perempuan", "TTL Ibu Perempuan", "Pekerjaan Ibu Perempuan", "Alamat Ibu Perempuan"
            ]),
            ("VI. DATA WALI, MAHAR & SAKSI", [
                "Nama Wali", "Bin Wali", "NIK Wali", "TTL Wali", "Pekerjaan Wali", "Alamat Wali", "Hubungan Wali", "Mahar / Maskawin",
                "Nama Saksi 1", "TTL Saksi 1", "NIK Saksi 1", "Pekerjaan Saksi 1", "Alamat Saksi 1",
                "Nama Saksi 2", "TTL Saksi 2", "NIK Saksi 2", "Pekerjaan Saksi 2", "Alamat Saksi 2"
            ])
        ]

        table_rows = [
            [
                Paragraph("<b>FIELD / PARAMETER</b>", cell_bold),
                Paragraph("<b>ISIAN DATA</b>", cell_bold)
            ]
        ]

        # Menyusun tabel tunggal penuh ke bawah untuk mode Portrait
        for title, keys in groups:
            table_rows.append([
                Paragraph(f"<b>{title}</b>", sec_header),
                Paragraph("", sec_header)
            ])
            for k in keys:
                val_txt = str(current_data.get(k, "") or "")
                table_rows.append([
                    Paragraph(f"<b>{k}</b>", cell_bold),
                    Paragraph(val_txt, cell_norm)
                ])

        col_widths = [200, 372]
        
        pdf_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
        
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F2C59")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 1.2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1.2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ]

        # Mewarnai header grup
        r_idx = 1
        for title, keys in groups:
            t_style.append(('BACKGROUND', (0, r_idx), (1, r_idx), colors.HexColor("#1E3A8A")))
            r_idx += len(keys) + 1

        pdf_table.setStyle(TableStyle(t_style))
        elements.append(pdf_table)
        elements.append(Spacer(1, 10))

        # BLOK TANDA TANGAN (Kasi Pelayanan & Kepala Desa Tambi)
        tgl_surat_str = current_data.get("Tanggal Surat", format_tanggal_indonesia(date.today()))
        
        ttd_rows = [
            [
                Paragraph("Petugas Pengantar / Kasi Pelayanan", ttd_center),
                Paragraph(f"Tambi, {tgl_surat_str}<br/>Mengetahui,<br/>Kepala Desa Tambi", ttd_center)
            ],
            [
                Paragraph("<br/><br/><br/>", ttd_center),
                Paragraph("<br/><br/><br/>", ttd_center)
            ],
            [
                Paragraph("<u><b>Chalim Muchtarom, S.Pd.I</b></u>", ttd_bold_underline),
                Paragraph("<u><b>J U R I</b></u>", ttd_bold_underline)
            ]
        ]
        
        ttd_table = Table(ttd_rows, colWidths=[286, 286])
        ttd_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))
        
        elements.append(ttd_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    pdf_bytes = generate_pdf_formal(st.session_state.get('input_data'))
    st.download_button(
        label="📄 Download Laporan PDF Ringkas Terkelompok (1 Lembar F4 Portrait)",
        data=pdf_bytes,
        file_name="Ringkasan_Data_Catin_F4.pdf",
        mime="application/pdf",
        use_container_width=True
    )
